import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# <Cell 'Sheet1'.A1>
# print(sheet['A1'])

# 2015-04-05 13:34:02
# print(sheet['A1'].value)

c = sheet['B1']
# Apples
# print(c.value)

# Row 1, Column 2 is Apples
# print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)

# 73
# print(sheet['C1'].value)

# Cell B1 is Apples
# print('Cell ' + c.coordinate + ' is ' + c.value)

# <Cell 'Sheet1'.B1>
# print(sheet.cell(row=1, column=2))

# Apples
# print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    # 1 Apples
    # 3 Pears
    # 5 Apples
    # 7 Strawberries
    print(i, sheet.cell(row=i, column=2).value)
