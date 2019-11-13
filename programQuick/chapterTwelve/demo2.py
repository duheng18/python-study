import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# 7
print(sheet.max_row)

# 3
print(sheet.max_column)
