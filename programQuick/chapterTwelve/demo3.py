import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# A
# print(get_column_letter(1))
# B
# print(get_column_letter(2))
# AA
# print(get_column_letter(27))
# AHP
# print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# 3
# print(sheet.max_column)

# C
print(get_column_letter(sheet.max_column))

# 1
print(column_index_from_string('A'))

# 27
print(column_index_from_string('AA'))
