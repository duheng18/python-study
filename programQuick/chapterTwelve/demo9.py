import openpyxl

wb = openpyxl.Workbook()
# ['Sheet']
print(wb.get_sheet_names())
# <Worksheet "Sheet1">
print(wb.create_sheet())
# <Worksheet "First Sheet">
print(wb.create_sheet(index=0, title='First Sheet'))
# ['First Sheet', 'Sheet', 'Sheet1']
print(wb.get_sheet_names())
# <Worksheet "Middle Sheet">
print(wb.create_sheet(index=2, title='Middle Sheet'))
# ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
# ['First Sheet', 'Sheet']
print(wb.get_sheet_names())
