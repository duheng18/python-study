'''

创建程序multiplicationTable.py，从命令行接受数字N，在一个Excel电子表格中创建一个N×N
的乘法表。例如，如果这样执行程序: py multiplicationTable.py 6
'''

import openpyxl, sys
from openpyxl.styles import Font

times = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb['Sheet']
boldFont = Font(bold=True)

for i in range(1, times + 1):
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = boldFont
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = boldFont

for i in range(2, times + 2):
    for j in range(2, times + 2):
        x = sheet.cell(row=i, column=1).value
        y = sheet.cell(row=1, column=j).value
        sheet.cell(row=i, column=j).value = x * y

wb.save('multiplicationTable.xlsx')
